from RPA.Browser.Selenium import Selenium
from RPA.FileSystem import FileSystem
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
import time

# All the variables required
url = "https://itdashboard.gov"
dest_filename = "output/output.xlsx"
agency_to_scrape = "Department of Labor"

# Empty objects are defined here for their global scope
browser = Selenium()
file_system = FileSystem()
browser.set_download_directory("output/downloads/")


def get_agencies_and_amounts():
    agencies_dict = {}
    browser.open_available_browser(url)
    browser.click_link("#home-dive-in")
    # browser.wait_until_element_is_visible("//span[contains(@class, 'w200')]")
    browser.wait_until_element_is_visible("//div[contains(@id, 'agency-tiles-container')]", 15)

    elem = browser.find_element("//div[contains(@id, 'agency-tiles-container')]")

    for key in elem.find_elements_by_class_name("tuck-5"):
        agency = key.find_element_by_class_name('w200').text
        amount = key.find_element_by_class_name('w900').text
        link = key.find_element_by_class_name('btn-sm')
        if agency not in agencies_dict.keys():
            agencies_dict.update({agency: {"amount": amount, "link": link}})

    return agencies_dict


def put_agencies_and_amounts_in_excel(agencies_dict):
    wb = Workbook()
    work_sheet = wb.create_sheet('Agencies')

    row = 2

    work_sheet.cell(row=1, column=1, value="Agencies")
    work_sheet.cell(row=1, column=2, value="Spendings")
    for key in agencies_dict:
        work_sheet.cell(row=row, column=1, value=key)
        work_sheet.cell(row=row, column=2, value=agencies_dict[key]["amount"])
        row += 1

    wb.save(filename=dest_filename)


def put_investments_in_excel(individual_investments):
    wb = load_workbook(dest_filename)
    work_sheet = wb.create_sheet('Individual Investments')

    row = 2

    work_sheet.cell(row=1, column=1, value="UII")
    work_sheet.cell(row=1, column=2, value="Bureau")
    work_sheet.cell(row=1, column=3, value="Investment Title")
    work_sheet.cell(row=1, column=4, value="Spendings")
    work_sheet.cell(row=1, column=5, value="Type")
    work_sheet.cell(row=1, column=6, value="Rating")
    work_sheet.cell(row=1, column=7, value="Projects")

    for key in individual_investments:
        work_sheet.cell(row=row, column=1, value=key)
        work_sheet.cell(row=row, column=2, value=individual_investments[key]["bureau"])
        work_sheet.cell(row=row, column=3, value=individual_investments[key]["title"])
        work_sheet.cell(row=row, column=4, value=individual_investments[key]["spending"])
        work_sheet.cell(row=row, column=5, value=individual_investments[key]["type"])
        work_sheet.cell(row=row, column=6, value=individual_investments[key]["rating"])
        work_sheet.cell(row=row, column=7, value=individual_investments[key]["projects"])
        row += 1

    wb.save(dest_filename)


def scrape_individual_investments(agencies_dict):
    individual_investments = {}
    browser.click_link(agencies_dict[agency_to_scrape]["link"])

    flag = True
    while flag:
        try:
            browser.wait_until_element_is_visible("//*[@id='investments-table-object']/tbody", 60)
            elem = browser.find_element("//*[@id='investments-table-object']/tbody")
            dropdown = browser.find_element('//*[@id="investments-table-object_length"]/label/select')
            browser.select_from_list_by_value(dropdown, "-1")
            browser.wait_until_element_is_visible("//*[@id='investments-table-object']/tbody/tr[11]", 30)
            for key in elem.find_elements_by_xpath("//*[@id='investments-table-object']/tbody/tr"):
                if key.text.split()[0] not in individual_investments:
                    uii = key.find_elements_by_css_selector("td")[0].text
                    if key.find_elements_by_css_selector("td")[0].get_attribute("innerHTML").__contains__("href"):
                        uii_link = key.find_elements_by_css_selector("td")[0].find_element_by_css_selector(
                            "a").get_attribute("href")
                    else:
                        uii_link = None
                    bureau = key.find_elements_by_css_selector("td")[1].text
                    title = key.find_elements_by_css_selector("td")[2].text
                    spending = key.find_elements_by_css_selector("td")[3].text
                    type = key.find_elements_by_css_selector("td")[4].text
                    rating = key.find_elements_by_css_selector("td")[5].text
                    projects = key.find_elements_by_css_selector("td")[6].text

                    individual_investments.update({uii: {
                        "uii": uii,
                        "uii_link": uii_link,
                        "bureau": bureau,
                        "title": title,
                        "spending": spending,
                        "type": type,
                        "rating": rating,
                        "projects": projects
                    }})
                # print(key.text)

            next_button = elem.find_element_by_xpath("//a[contains(@class, 'next')]")

            if "disabled" in next_button.get_attribute("class"):
                flag = False
            else:
                browser.click_button(next_button)

        except (StaleElementReferenceException, NoSuchElementException):
            continue
    return individual_investments


def download_pdfs(individual_investments):
    for key in individual_investments:
        if individual_investments[key]["uii_link"]:
            browser.go_to(individual_investments[key]["uii_link"])
            browser.wait_until_element_is_visible('//*[@id="business-case-pdf"]/a')
            browser.click_link(browser.find_element('//*[@id="business-case-pdf"]/a'))
            time.sleep(10)



def main():
    try:
        agencies_dict = get_agencies_and_amounts()
        put_agencies_and_amounts_in_excel(agencies_dict)
        individual_investments = scrape_individual_investments(agencies_dict)
        put_investments_in_excel(individual_investments)
        download_pdfs(individual_investments)
    finally:
        browser.close_browser()



